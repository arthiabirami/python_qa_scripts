import openpyxl

#file-->workbook-->sheets-->rows-->cells

file = "/Users/sathish/code/python testing/data1.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet 1"]

rows=sheet.max_row #count no of rows in excel sheet
columns=sheet.max_column #count no of columns in excel sheet

#reading all the rows and columns 
for r in range (1,rows+1):
	for c in range (1,columns+1):
		val=(sheet.cell(r,c).value) # read the data in the sheet
		print (val)
	if val == 'None':
		break;
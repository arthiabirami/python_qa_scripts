import openpyxl

file = "/Users/sathish/code/python testing/sample1.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active

# for r in range (1,6):
# 	for c in range (1,4):
# 		sheet.cell(r,c).value= "welcome"   # write  the data in the sheet

sheet.cell(1,1).value=123
sheet.cell(1,2).value="smith"
sheet.cell(1,3).value="doctor"

sheet.cell(2,1).value=456
sheet.cell(2,2).value="don"
sheet.cell(2,3).value="doctor"

sheet.cell(3,1).value=789
sheet.cell(3,2).value="roger"
sheet.cell(3,3).value="doctor"

workbook.save(file)

